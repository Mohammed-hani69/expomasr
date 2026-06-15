import sqlite3
import os
import csv
import io
from datetime import datetime
from functools import wraps
from flask import Flask, request, jsonify, g, render_template, redirect, url_for, session, Response
from flask_cors import CORS

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'expo-masr-2026-secret-key-change-in-production')
# Enable CORS for all routes
CORS(app, resources={r"/*": {"origins": "*", "methods": ["GET", "POST", "OPTIONS"], "allow_headers": ["Content-Type"]}})

DB_PATH = os.path.join(os.path.dirname(__file__), 'bookings.db')

# Admin credentials (change in production via env vars)
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'ezezo291')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'ezezo291')


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


def close_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS bookings (
            id TEXT PRIMARY KEY,
            companyName TEXT NOT NULL,
            contactPerson TEXT NOT NULL,
            phone TEXT NOT NULL,
            whatsapp TEXT NOT NULL,
            email TEXT DEFAULT '',
            city TEXT DEFAULT '',
            sector TEXT NOT NULL,
            selectedPackage TEXT NOT NULL,
            price REAL DEFAULT 0,
            message TEXT DEFAULT '',
            date TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


@app.teardown_appcontext
def teardown(exception=None):
    close_db(exception)


# ─── API Routes ────────────────────────────────────────────────

@app.route('/api/bookings', methods=['POST'])
def create_booking():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'بيانات غير صالحة'}), 400

    required = ['id', 'companyName', 'contactPerson', 'phone', 'whatsapp', 'sector', 'selectedPackage']
    for field in required:
        if not data.get(field):
            return jsonify({'success': False, 'error': f'الحقل {field} مطلوب'}), 400

    try:
        db = get_db()
        db.execute('''
            INSERT INTO bookings (id, companyName, contactPerson, phone, whatsapp, email, city, sector, selectedPackage, price, message, date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['id'],
            data['companyName'],
            data['contactPerson'],
            data['phone'],
            data['whatsapp'],
            data.get('email', ''),
            data.get('city', ''),
            data['sector'],
            data['selectedPackage'],
            data.get('price', 0),
            data.get('message', ''),
            data.get('date', datetime.now().isoformat()),
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم حفظ الحجز بنجاح', 'id': data['id']}), 201
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'كود الحجز موجود مسبقاً'}), 409
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bookings', methods=['GET'])
def get_bookings():
    try:
        db = get_db()
        rows = db.execute('SELECT * FROM bookings ORDER BY created_at DESC').fetchall()
        bookings = [dict(row) for row in rows]
        return jsonify({'success': True, 'data': bookings, 'count': len(bookings)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bookings/stats', methods=['GET'])
def get_stats():
    try:
        db = get_db()
        total = db.execute('SELECT COUNT(*) as count FROM bookings').fetchone()['count']
        return jsonify({'success': True, 'total': total})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ─── Admin Dashboard Routes ────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            error = 'اسم المستخدم أو كلمة المرور غير صحيحة'
    return render_template('admin/login.html', error=error)


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))


@app.route('/admin/')
@login_required
def admin_dashboard():
    db = get_db()
    search = request.args.get('search', '').strip()
    sector_filter = request.args.get('sector', '').strip()

    query = 'SELECT * FROM bookings WHERE 1=1'
    params = []

    if search:
        query += ' AND (companyName LIKE ? OR contactPerson LIKE ? OR phone LIKE ? OR id LIKE ?)'
        like = f'%{search}%'
        params.extend([like, like, like, like])

    if sector_filter:
        query += ' AND sector = ?'
        params.append(sector_filter)

    query += ' ORDER BY created_at DESC'
    rows = db.execute(query, params).fetchall()
    bookings = [dict(row) for row in rows]

    # Stats
    total = db.execute('SELECT COUNT(*) as count FROM bookings').fetchone()['count']
    last = db.execute('SELECT companyName FROM bookings ORDER BY created_at DESC LIMIT 1').fetchone()
    premium_count = db.execute("SELECT COUNT(*) as count FROM bookings WHERE selectedPackage LIKE '%Premium%' OR selectedPackage LIKE '%Sponsor%' OR price >= 12000").fetchone()['count']

    # Unique sectors for filter dropdown
    sectors = [r['sector'] for r in db.execute('SELECT DISTINCT sector FROM bookings ORDER BY sector').fetchall()]

    return render_template('admin/dashboard.html',
        bookings=bookings,
        search=search,
        sector=sector_filter,
        sectors=sectors,
        stats={'total': total, 'last_booking': last['companyName'] if last else None, 'premium_count': premium_count}
    )


@app.route('/admin/export')
@login_required
def admin_export():
    """تصدير البيانات بصيغة CSV"""
    db = get_db()
    rows = db.execute('SELECT * FROM bookings ORDER BY created_at DESC').fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['كود الحجز', 'اسم الشركة', 'المسؤول', 'الهاتف', 'واتساب', 'البريد', 'المدينة', 'القطاع', 'الباقة', 'السعر', 'الرسالة', 'تاريخ الحجز'])
    for r in rows:
        writer.writerow([r['id'], r['companyName'], r['contactPerson'], r['phone'], r['whatsapp'], r['email'], r['city'], r['sector'], r['selectedPackage'], r['price'], r['message'], r['date']])

    csv_bytes = output.getvalue().encode('utf-8-sig')
    return Response(
        csv_bytes,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=expo_masr_bookings_{datetime.now().strftime("%Y%m%d")}.csv'}
    )


@app.route('/api/generate-ticket-pdf', methods=['POST'])
def generate_ticket_pdf():
    """توليد PDF للتذكرة من الـ backend"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'بيانات غير صالحة'}), 400
        
        # Extract booking data
        booking_id = data.get('id', 'REG-2026')
        company_name = data.get('companyName', 'شركة')
        package = data.get('selectedPackage', 'حزمة قياسية')
        price = data.get('price', 0)
        contact_person = data.get('contactPerson', 'المسؤول')
        phone = data.get('phone', '---')
        whatsapp = data.get('whatsapp', '---')
        
        # Create PDF in memory
        output = io.BytesIO()
        
        # Setup PDF with A4 size in portrait mode
        pdf_width, pdf_height = A4
        c = canvas.Canvas(output, pagesize=A4)
        
        # Set colors
        dark_blue = (3/255, 11/255, 26/255)  # #030b1a
        gold = (212/255, 175/255, 55/255)    # #d4af37
        
        # Draw dark blue background
        c.setFillColor(*dark_blue)
        c.rect(0, 0, pdf_width, pdf_height, fill=1, stroke=0)
        
        # Draw gold borders
        c.setLineWidth(2)
        c.setStrokeColor(*gold)
        c.rect(1*cm, 1*cm, pdf_width-2*cm, pdf_height-2*cm)
        
        c.setLineWidth(0.5)
        c.rect(0.7*cm, 0.7*cm, pdf_width-1.4*cm, pdf_height-1.4*cm)
        
        # Title
        c.setFont("Helvetica-Bold", 28)
        c.setFillColor(*gold)
        c.drawCentredString(pdf_width/2, pdf_height - 2*cm, "معرض مصر 2026")
        
        # Subtitle
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(1, 1, 1)
        c.drawCentredString(pdf_width/2, pdf_height - 2.5*cm, "رخصة وتأكيد الجناح")
        
        # Draw a line separator
        c.setLineWidth(1)
        c.setStrokeColor(*gold)
        c.line(1.5*cm, pdf_height - 2.8*cm, pdf_width - 1.5*cm, pdf_height - 2.8*cm)
        
        # Ticket details
        y_position = pdf_height - 3.5*cm
        line_height = 0.6*cm
        
        # Company info section
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(*gold)
        c.drawString(1.5*cm, y_position, "بيانات الشركة")
        y_position -= line_height
        
        c.setFont("Helvetica", 10)
        c.setFillColor(1, 1, 1)
        
        # Company name
        c.drawString(2*cm, y_position, f"اسم الشركة: {company_name}")
        y_position -= line_height
        
        # Contact person
        c.drawString(2*cm, y_position, f"المسؤول: {contact_person}")
        y_position -= line_height
        
        # Phone & WhatsApp on same line
        c.drawString(2*cm, y_position, f"الهاتف: {phone}")
        c.drawString(4.5*cm, y_position, f"واتساب: {whatsapp}")
        y_position -= line_height * 1.5
        
        # Package info section
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(*gold)
        c.drawString(1.5*cm, y_position, "تفاصيل الحجز")
        y_position -= line_height
        
        c.setFont("Helvetica", 10)
        c.setFillColor(1, 1, 1)
        
        # Package
        c.drawString(2*cm, y_position, f"الباقة: {package}")
        y_position -= line_height
        
        # Price
        c.drawString(2*cm, y_position, f"السعر: {price:,.2f} ج.م")
        y_position -= line_height
        
        # Booking ID
        c.drawString(2*cm, y_position, f"رقم الحجز: {booking_id}")
        y_position -= line_height * 1.5
        
        # Important notes section
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(*gold)
        c.drawString(1.5*cm, y_position, "ملاحظات مهمة:")
        y_position -= line_height
        
        c.setFont("Helvetica", 9)
        c.setFillColor(1, 1, 1)
        
        notes = [
            "✓ يرجى الاحتفاظ بهذه الرخصة وإحضارها يوم المعرض",
            "✓ تأكد من صحة جميع البيانات المدرجة أعلاه",
            "✓ للاستفسارات يرجى التواصل عبر واتساب أو الهاتف",
        ]
        
        for note in notes:
            c.drawString(2*cm, y_position, note)
            y_position -= line_height
        
        # Footer
        y_position = 1.5*cm
        c.setFont("Helvetica", 8)
        c.setFillColor(0.7, 0.7, 0.7)
        c.drawCentredString(pdf_width/2, y_position, f"تم إنشاء هذه الرخصة في {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Finalize PDF
        c.save()
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=expo_masr_ticket_{booking_id}.pdf'}
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/export-excel')
@login_required
def admin_export_excel():
    """تصدير البيانات بصيغة Excel احترافية مع تنسيق مميز"""
    if not HAS_OPENPYXL:
        return jsonify({'success': False, 'error': 'مكتبة Excel غير مثبتة'}), 500
    
    db = get_db()
    rows = db.execute('SELECT * FROM bookings ORDER BY created_at DESC').fetchall()
    
    # إنشاء workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "الحجوزات"
    
    # تعيين عرض الأعمدة
    column_widths = [15, 20, 15, 15, 15, 20, 15, 20, 20, 15, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # تنسيق الرأس
    headers = ['كود الحجز', 'اسم الشركة', 'المسؤول', 'الهاتف', 'واتساب', 'البريد', 'المدينة', 'القطاع', 'الباقة', 'السعر', 'الرسالة', 'تاريخ الحجز']
    
    # ألوان الذهب والأزرق (ألوان المعرض)
    gold_fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
    gold_font = Font(bold=True, color='030B1A', size=12, name='Calibri')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # كتابة رؤوس الأعمدة
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = gold_fill
        cell.font = gold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # تنسيق الصفوف
    data_font = Font(size=11, name='Calibri')
    alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    for row_idx, row_data in enumerate(rows, 2):
        # تلوين الصفوف بشكل متناوب
        row_fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        
        values = [
            row_data['id'],
            row_data['companyName'],
            row_data['contactPerson'],
            row_data['phone'],
            row_data['whatsapp'],
            row_data['email'],
            row_data['city'],
            row_data['sector'],
            row_data['selectedPackage'],
            f"{row_data['price']:,.2f}",
            row_data['message'],
            row_data['date']
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col in [1, 8] else 'left', vertical='center', wrap_text=True)
            cell.fill = row_fill
    
    # إنشاء ملف في الذاكرة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=expo_masr_bookings_{datetime.now().strftime("%Y%m%d")}.xlsx'}
    )


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=True)
